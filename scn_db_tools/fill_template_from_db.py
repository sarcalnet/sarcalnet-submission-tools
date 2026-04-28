# The MIT License (MIT)
# Copyright (c) 2025 by the xcube team
#
# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including without limitation
# the rights to use, copy, modify, merge, publish, distribute, sublicense,
# and/or sell copies of the Software, and to permit persons to whom the
# Software is furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
# FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.
import warnings
from typing import List, Tuple

import click
import pandas as pd
from geopandas import GeoDataFrame
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from pandas import ExcelWriter
from xcube_geodb.core.geodb import GeoDBClient


class DataFetcher:
    def __init__(
        self,
        geodb: GeoDBClient,
    ):
        self._geodb = geodb

    def fetch_data(
        self, unique_site_ids: List[str]
    ) -> dict[str, None] | dict[str, GeoDataFrame]:
        short_site_ids = [i[:4] for i in unique_site_ids]
        primary_tts = [i[5:] for i in unique_site_ids]
        sites_where = ""
        for i, ssid in enumerate(short_site_ids):
            sites_where += f"(short_site_id = '{ssid}' AND primary_target_type_id like '{primary_tts[i]}%')"
            sites_where += " OR "
        sites_where = sites_where[:-4]

        sites = self._geodb.get_collection_pg(
            collection="calibration_sites", where=sites_where, database="sarcalnet"
        )

        if sites is None or sites.empty:
            return {
                "sites": None,
                "targets": None,
                "surveys": None,
                "nat_targets": None,
                "nat_surveys": None,
            }

        targets_where = ""
        for row in sites.iterrows():
            site_name = row[1]["site_name"]
            targets_where += f"site_name = '{site_name}'"
            targets_where += " OR "
        targets_where = targets_where[:-4]

        targets = self._geodb.get_collection_pg(
            collection="calibration_targets", where=targets_where, database="sarcalnet"
        )

        nat_targets = self._geodb.get_collection_pg(
            collection="calibration_nat_targets",
            where=targets_where,
            database="sarcalnet",
        )

        if targets is not None:
            surveys_where = ""
            for row in targets.iterrows():
                target_id = row[1]["target_id"]
                surveys_where += f"target_id = '{target_id}'"
                surveys_where += " OR "
            surveys_where = surveys_where[:-4]

            surveys = self._geodb.get_collection_pg(
                collection="calibration_surveys",
                where=surveys_where,
                database="sarcalnet",
            )
        else:
            surveys = None

        if nat_targets is not None and not nat_targets.empty:
            nat_surveys_where = ""
            for row in nat_targets.iterrows():
                target_id = row[1]["target_id"]
                nat_surveys_where += f"target_id = '{target_id}'"
                nat_surveys_where += " OR "
            nat_surveys_where = nat_surveys_where[:-4]

            nat_surveys = self._geodb.get_collection_pg(
                collection="calibration_nat_surveys",
                where=nat_surveys_where,
                database="sarcalnet",
            )
        else:
            nat_surveys = None

        return {
            "sites": sites,
            "targets": targets,
            "surveys": surveys,
            "nat_targets": nat_targets
            if nat_targets is not None and not nat_targets.empty
            else None,
            "nat_surveys": nat_surveys
            if nat_surveys is not None and not nat_surveys.empty
            else None,
        }


class Outputter:

    @staticmethod
    def copy_sheet_formatting(source_file, sheet_name, target_file):
        """
        Copy formatting from one sheet in source_file to a sheet in target_file,
        keeping the target values unchanged.

        Parameters
        ----------
        source_file : str
            Path to the source Excel file (contains formatting to copy).
        sheet_name : str
            Name of the sheet in the source file to copy formatting from.
        target_file : str
            Path to the target Excel file (contains values to keep).
        """
        # Load workbooks
        wb_source = load_workbook(source_file)
        wb_target = load_workbook(target_file)

        if not sheet_name in wb_source.sheetnames:
            return

        source = wb_source[sheet_name]
        target = wb_target[sheet_name]

        for row in source.iter_rows():
            for cell in row:
                target_cell = target.cell(row=cell.row, column=cell.column)

                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = copy(cell.number_format)
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)

        if sheet_name in ["site", "cr", "dt", "survey"]:
            fixed_width = 43
            for col_idx in range(1, target.max_column + 1):
                col_letter = get_column_letter(col_idx)
                if col_letter in target.column_dimensions:
                    del target.column_dimensions[col_letter]  # remove old dimension
                target.column_dimensions[col_letter].width = fixed_width
        else:
            max_cols = max(source.max_column, target.max_column)

            for col_idx in range(1, max_cols + 1):
                source_letter = get_column_letter(col_idx)
                target_letter = get_column_letter(col_idx)

                if source_letter in source.column_dimensions:
                    source_width = source.column_dimensions[source_letter].width
                    if source_width is not None:
                        target.column_dimensions[target_letter].width = source_width

        # Copy merged cells
        for merged in source.merged_cells.ranges:
            try:
                target.merge_cells(str(merged))
            except ValueError:
                # skip if the merged range doesn't fit target sheet
                pass

        wb_target.save(target_file)


    def write_form(
        self, source_file: str, data: dict[str, None] | dict[str, GeoDataFrame], target_file: str
    ):
        source = pd.read_excel(source_file, sheet_name=None)

        with ExcelWriter(target_file, "openpyxl") as excel_writer:
            readme = source["README"].copy()
            readme.columns = [""] * len(readme.columns)
            readme.to_excel(
                excel_writer, sheet_name="README", index=False, header=True
            )

            sites: GeoDataFrame = data["sites"]

            sites["unique_site_id"] = (
                sites["short_site_id"]
                + "-"
                + sites["primary_target_type_id"].astype(str).str[:2]
            )
            sites["special_requests"] = sites["special_requests"].replace({True: "Y", False: "N"})
            sites.insert(0, "unique_site_id", sites.pop("unique_site_id"))
            sites.rename(
                columns={
                    "unique_site_id": "Unique Site ID",
                    "short_site_id": "Short Site ID",
                    "country": "Country",
                    "site_name": "Site Name",
                    "province_state_region": "Province / state / region",
                    "primary_target_type_id": "Primary Target Type ID",
                    "target_types": "Target Types",
                    "primary_sensor": "Primary Sensor",
                    "special_requests": "Willing to consider special requests",
                    "responsible_organization": "Responsible Organization",
                    "website": "Website",
                    "active_from": "Active from  (YYYY-MM-DD)",
                    "active_until": 'Active until (YYYY-MM-DD or "-")',
                    "poc_name": "POC Name",
                    "poc_email": "POC email",
                    "poc_name2": "Additional POC Name",
                    "poc_email2": "Additional POC email",
                    "centroid": "Centroid of the site (latitude and longitude in decima deg)",
                    "boundaries": "Boundaries",
                    "maintenance_schedule": "Planned maintenance schedule",
                    "landcover": "Characteristics",
                },
                inplace=True,
            )

            source_sites = source["site"].iloc[:4]
            sites_merged = pd.concat([source_sites, sites.reindex(columns=source_sites.columns)], ignore_index=True)

            sites_merged.to_excel(
                excel_writer, sheet_name="site", index=False
            )
            targets: GeoDataFrame = data["targets"]
            if len(targets) > 0:
                targets["operational"] = targets["operational"].replace({True: "Y", False: "N"})
                targets["rcs_angle_dependency_availability"] = targets["rcs_angle_dependency_availability"].replace({True: "Y", False: "N"})
                targets.rename(
                    columns={
                        "target_id": "Unique Target ID",
                        "short_site_id": "Short Site ID",
                        "site_name": "Site Name",
                        "subsite": "Sub-site",
                        "internal_id": "Internal ID",
                        "short_target_id": "Short Target ID",
                        "target_type": "Target Type ID",
                        "target_description": "Target description",
                        "approx_lat": "Approximate Latitude\n(decimal deg, WGS84)",
                        "approx_lon": "Approximate Longitude\n(decimal deg WGS84)",
                        "approx_h": "Approximate elevation\n(meters, WGS84)",
                        "approx_azimuth_angle": "Approximate Boresight Azimuth angle\n(decimal deg)",
                        "approx_boresight_angle": "Approximate Boresight Elevation angle\n(decimal deg)",
                        "primary_direction": "Primary direction",
                        "side_length": "Side length (m)",
                        "photo_link": "Photo link",
                        "operational": "Operational",
                        "manufacturer": "Manufacturer",
                        "purpose": "Purpose of target",
                        "rcs": "Reference RCS (dBm2)",
                        "rcs_measurement_conditions": "Reference RCS measurement sensor",
                        "reference_rcs_accuracy": "Reference RCS measurement expected accuracy (dB)",
                        "reference_rcs_boresight_angle": "Reference RCS measurement boresite angle (decimal deg)",
                        "reference_rcs_wavelength": "Reference RCS measurement wavelength (m)",
                        "reference_rcs_bandwidth": "Reference RCS measurement bandwidth (Hz)",
                        "rcs_accuracy_determination": "RCS accuracy determination method",
                        "rcs_angle_dependency_availability": "RCS angle dependency availablity",
                        "composition": "Composition",
                        "characterization": "Characterization of reflector",
                    },
                    inplace=True,
                )
                source_targets = source["cr"].iloc[:4]
                targets = pd.concat([source_targets, targets.reindex(columns=source_targets.columns)], ignore_index=True)
                targets.to_excel(
                    excel_writer, sheet_name="cr", index=False
                )

            surveys: GeoDataFrame = data["surveys"]
            if len(surveys) > 0:
                surveys["fence"] = surveys["fence"].replace({True: "Y", False: "N"})
                surveys.rename(
                    columns={
                        "target_id": "Unique Target ID",
                        "survey_date": "Survey date (YYYY-MM-DD)",
                        "lat": "Latitude (decimal deg)",
                        "lon": "Longitude (decimal deg)",
                        "elevation": "Elevation (m)",
                        "position_accuracy": "Position accuracy (cm)",
                        "crs": "Coordinate Reference System (WKT or EPSG)",
                        "crs_vx": "CRS X velocity (mm/year)",
                        "crs_vy": "CRS Y velocity (mm/year)",
                        "crs_vz": "CRS Z velocity (mm/year)",
                        "azimuth_angle": "Boresight azimuth angle\n(decimal deg)",
                        "boresight_angle": "Boresight elevation angle\n(decimal deg)",
                        "tilt_angle": "Tilt (decimal deg)",
                        "pointing_accuracy": "Pointing accuracy\n(decimal deg)",
                        "fence": "Fence",
                        "measurement_method": "Measurement method",
                        "offset_method": "Offset method",
                        "applied_corrections": "Applied corrections",
                        "gnss_measurement_duration": "GNSS measurement duration\n(hh:mm:ss)",
                        "photo_link": "Photo link",
                        "report_status": "Status report",
                    },
                    inplace=True,
                )

                source_surveys = source["survey"].iloc[:4]
                surveys = pd.concat([source_surveys, surveys.reindex(columns=source_surveys.columns)], ignore_index=True)
                surveys.to_excel(
                    excel_writer, sheet_name="survey", index=False
                )
            nat_targets: GeoDataFrame = data["nat_targets"]
            if nat_targets is not None and "dt" in source.keys():
                nat_targets.rename(
                    columns={
                        "target_id": "Unique Target ID",
                        "short_site_id": "Short Site ID",
                        "site_name": "Site Name",
                        "subsite": "Sub-site",
                        "internal_id": "Internal ID",
                        "short_target_id": "Short Target ID",
                        "target_type": "Target Type ID",
                        "target_description": "Target description",
                        "geometry": "Bounding polygon (WKT, WGS84)",
                        "coverage": "Coverage (km2)",
                        "mask_polygon": "Mask polygon (WKT, WGS84)",
                        "start_monitoring": "Start Monitoring Period (YYYY-MM-DD)",
                        "stop_monitoring": 'Stop Monitoring Period (YYYY-MM-DD or "-")',
                    },
                    inplace=True,
                )
                source_nat_targets = source["dt"].iloc[:4]
                nat_targets = pd.concat([source_nat_targets, nat_targets.reindex(columns=source_nat_targets.columns)], ignore_index=True)
                nat_targets.to_excel(
                    excel_writer, sheet_name="dt", index=False
                )
            nat_surveys: GeoDataFrame = data["nat_surveys"]
            if nat_surveys is not None:
                nat_surveys.rename(
                    columns={
                        "target_id": "Unique Target ID",
                        "survey_start": "Start Survey Period (YYYY-MM-DD)",
                        "survey_stop": "Stop Survey Period (YYYY-MM-DD)",
                        "mission": "Mission",
                        "carrier_frequency": "Carrier Frequency (GHz)",
                        "polarizations": "Polarization Channel",
                        "observation_time": "UTC Observation Time (HH:MM)",
                        "local_observation_time": "Local Observation time (HH:MM)",
                        "incidence_angle_range": "Incidence Angle Range (min - max, in decimal deg)",
                        "backscatter_coefficient_type": "Backscatter coefficient type",
                        "backscatter_coeff_mean": "Mean Backscatter Coefficient (dB)",
                        "backscatter_coeff_std": "Backscatter Coefficient Standard Deviation (dB)",
                        "reference_surface": "Reference Surface",
                        "samples": "Samples",
                        "relative_orbit": "Relative Orbit",
                        "orbit_dir": "Orbit direction",
                        "look_side": "Look side",
                        "acquisition_mode": "Acquisition Mode",
                        "beam_id": "Beam ID",
                        "scene_ids": "Scene identifier(s)",
                        "query_url": "Query URL",
                    },
                    inplace=True,
                )

                source_nat_surveys = source["survey"].iloc[:4]
                nat_surveys = pd.concat([source_nat_surveys, nat_surveys.reindex(columns=source_nat_surveys.columns)], ignore_index=True)
                nat_surveys.to_excel(
                    excel_writer, sheet_name="dt", index=False
                )

                nat_surveys.to_excel(
                    excel_writer,
                    sheet_name="survey",
                    index=False,
                )


            if "unavailability (optional)" in source.keys():
                unavailability = source["unavailability (optional)"].copy()
                unavailability.to_excel(
                    excel_writer, sheet_name="unavailability (optional)", index=False
                )

            if "definitions" in source.keys():
                definitions = source["definitions"].copy()
                definitions.columns = [""] * len(definitions.columns)
                source["definitions"].to_excel(
                    excel_writer, sheet_name="definitions", index=False, header=False
                )
        self.copy_sheet_formatting(source_file, "README", target_file)
        self.copy_sheet_formatting(source_file, "site", target_file)
        self.copy_sheet_formatting(source_file, "cr", target_file)
        self.copy_sheet_formatting(source_file, "dt", target_file)
        self.copy_sheet_formatting(source_file, "survey", target_file)
        self.copy_sheet_formatting(source_file, "unavailability (optional)", target_file)
        self.copy_sheet_formatting(source_file, "definitions", target_file)


@click.command()
@click.argument(
    "source_file",
    metavar="SOURCE_FILE",
)
@click.argument(
    "target_file",
    metavar="TARGET_FILE",
)
@click.option("--client_id", required=True, help="The geoDB client_id.")
@click.option("--client_secret", required=True, help="The geoDB client_secret.")
@click.option(
    "--server_url",
    default="https://xcube-geodb.brockmann-consult.de",
    help="The geoDB server URL.",
)
@click.option("--server_port", default=443, help="The geoDB server port.")
@click.option(
    "--auth_domain",
    default="https://winchester.production.brockmann-consult.de/winchester",
    help="The geoDB auth domain URL.",
)
@click.argument(
    "unique_site_ids",
    metavar="UNIQUE_SITE_IDS",
    nargs=-1
)
def create_form(
    source_file: str,
    target_file: str,
    client_id: str,
    client_secret: str,
    server_url: str,
    server_port: int,
    auth_domain: str,
    unique_site_ids: Tuple[str],
):
    """
    Creates a filled template from the information of the SARCalNet database.
    """

    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    geodb = GeoDBClient(
        server_url, server_port, client_id, client_secret, auth_domain=auth_domain
    )
    data = DataFetcher(geodb).fetch_data(list(unique_site_ids))
    if not target_file.endswith(".xlsx"):
        target_file = target_file + ".xlsx"
    Outputter().write_form(source_file, data, target_file)
    print(f"Written filled template to {target_file}.")


if __name__ == "__main__":
    create_form()
